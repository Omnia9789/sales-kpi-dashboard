"""
reports/report_generator.py — Auto-generates a formatted Excel KPI report.

Run directly:
    python reports/report_generator.py

Or call from main.py / Streamlit dashboard.
"""

import os
import sys
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
from config import CLEAN_CSV, REPORTS_OUTPUT_DIR, TOP_N, MOM_DROP_THRESHOLD
from analysis.trend_detector import compute_monthly_revenue, detect_drops


# ── Style helpers ─────────────────────────────────────────────────────────────
DARK_BLUE  = "1F3864"
MID_BLUE   = "2E75B6"
LIGHT_BLUE = "BDD7EE"
RED        = "FF0000"
GREEN      = "70AD47"
YELLOW     = "FFD700"
WHITE      = "FFFFFF"
LIGHT_GREY = "F2F2F2"


def _header_fill(color: str = DARK_BLUE) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _font(bold=False, color=WHITE, size=11) -> Font:
    return Font(bold=bold, color=color, size=size, name="Calibri")


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header_row(ws, row: int, n_cols: int,
                      fill_color: str = DARK_BLUE) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill    = _header_fill(fill_color)
        cell.font    = _font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border  = _thin_border()


def _style_data_row(ws, row: int, n_cols: int, alt: bool = False) -> None:
    fill = PatternFill("solid", fgColor=LIGHT_GREY if alt else WHITE)
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill   = fill
        cell.font   = Font(name="Calibri", size=10, color="000000")
        cell.border = _thin_border()
        cell.alignment = Alignment(vertical="center")


def _auto_width(ws) -> None:
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) if cell.value is not None else 0)
            for cell in col
        )
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)


def _write_df_to_sheet(ws, df: pd.DataFrame, start_row: int = 1,
                       header_color: str = DARK_BLUE) -> int:
    """Write a DataFrame into a worksheet; return the next empty row."""
    # Header
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=col_idx, value=str(col_name).replace("_", " ").title())
    _style_header_row(ws, start_row, len(df.columns), header_color)

    # Data rows
    for r_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
        _style_data_row(ws, r_idx, len(df.columns), alt=(r_idx % 2 == 0))

    _auto_width(ws)
    return start_row + len(df) + 2


# ── Sheet builders ─────────────────────────────────────────────────────────────

def _build_summary_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Executive Summary")
    ws.sheet_view.showGridLines = False

    # Title block
    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value     = "Sales KPI Dashboard — Executive Summary"
    title_cell.font      = Font(bold=True, size=16, color=WHITE, name="Calibri")
    title_cell.fill      = _header_fill(DARK_BLUE)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    ts = ws["A2"]
    ts.value     = f"Generated: {datetime.now().strftime('%Y-%m-%d  %H:%M')}"
    ts.font      = Font(size=10, color="555555", italic=True, name="Calibri")
    ts.alignment = Alignment(horizontal="center")

    # KPI cards (row 4)
    kpis = [
        ("Total Revenue",    f"${df['sales'].sum():,.0f}",   MID_BLUE),
        ("Total Profit",     f"${df['profit'].sum():,.0f}",  GREEN),
        ("Avg Profit Margin",f"{(df['profit'].sum()/df['sales'].sum()*100):.1f}%", YELLOW),
        ("Total Orders",     f"{df['order_id'].nunique():,}", DARK_BLUE),
        ("Unique Customers", f"{df['customer_id'].nunique():,}", MID_BLUE),
        ("Unique Products",  f"{df['product_id'].nunique():,}", GREEN),
    ]
    for i, (label, value, color) in enumerate(kpis, 1):
        col = get_column_letter(i)
        ws.merge_cells(f"{col}4:{col}5")
        ws.merge_cells(f"{col}6:{col}7")
        lbl_cell = ws[f"{col}4"]
        val_cell = ws[f"{col}6"]
        lbl_cell.value     = label
        lbl_cell.fill      = _header_fill(color)
        lbl_cell.font      = Font(bold=True, size=11, color=WHITE, name="Calibri")
        lbl_cell.alignment = Alignment(horizontal="center", vertical="center")
        val_cell.value     = value
        val_cell.fill      = PatternFill("solid", fgColor=LIGHT_BLUE)
        val_cell.font      = Font(bold=True, size=14, name="Calibri")
        val_cell.alignment = Alignment(horizontal="center", vertical="center")


def _build_revenue_by_region(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Revenue by Region")
    ws.sheet_view.showGridLines = False

    region_df = (
        df.groupby("region")
        .agg(
            total_revenue=("sales",  "sum"),
            total_profit =("profit", "sum"),
            order_count  =("order_id","nunique"),
        )
        .round(2)
        .reset_index()
        .sort_values("total_revenue", ascending=False)
    )
    region_df["profit_margin_pct"] = (
        region_df["total_profit"] / region_df["total_revenue"] * 100
    ).round(2)

    _write_df_to_sheet(ws, region_df, start_row=2, header_color=MID_BLUE)

    ws.merge_cells("A1:E1")
    h = ws["A1"]
    h.value     = "Revenue & Profit by Region"
    h.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    h.fill      = _header_fill(DARK_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")


def _build_top_products(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(f"Top {TOP_N} Products")
    ws.sheet_view.showGridLines = False

    prod_df = (
        df.groupby(["product_name", "category", "sub_category"])
        .agg(total_revenue=("sales","sum"), total_profit=("profit","sum"),
             units_sold=("quantity","sum"))
        .round(2)
        .reset_index()
        .sort_values("total_revenue", ascending=False)
        .head(TOP_N)
    )

    _write_df_to_sheet(ws, prod_df, start_row=2, header_color=MID_BLUE)

    ws.merge_cells("A1:F1")
    h = ws["A1"]
    h.value     = f"Top {TOP_N} Products by Revenue"
    h.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    h.fill      = _header_fill(DARK_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")


def _build_category_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Category Breakdown")
    ws.sheet_view.showGridLines = False

    cat_df = (
        df.groupby(["category","sub_category"])
        .agg(total_revenue=("sales","sum"), total_profit=("profit","sum"),
             avg_discount=("discount","mean"))
        .round(2)
        .reset_index()
        .sort_values(["category","total_revenue"], ascending=[True, False])
    )
    cat_df["profit_margin_pct"] = (
        cat_df["total_profit"] / cat_df["total_revenue"] * 100
    ).round(2)

    _write_df_to_sheet(ws, cat_df, start_row=2, header_color=MID_BLUE)

    ws.merge_cells("A1:F1")
    h = ws["A1"]
    h.value     = "Revenue by Category & Sub-Category"
    h.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    h.fill      = _header_fill(DARK_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")


def _build_monthly_trend(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Monthly Trend")
    ws.sheet_view.showGridLines = False

    monthly = compute_monthly_revenue()
    monthly = detect_drops(monthly)

    # Colour-code drops
    next_row = _write_df_to_sheet(ws, monthly, start_row=2, header_color=MID_BLUE)

    # Highlight drop rows in red
    for r_idx, row in enumerate(monthly.itertuples(index=False), 3):
        if row.is_drop:
            for col in range(1, len(monthly.columns) + 1):
                ws.cell(row=r_idx, column=col).fill = PatternFill("solid", fgColor="FFCCCC")

    ws.merge_cells("A1:G1")
    h = ws["A1"]
    h.value     = f"Monthly Revenue Trend  (⚠️ red = drop > {MOM_DROP_THRESHOLD*100:.0f}%)"
    h.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    h.fill      = _header_fill(DARK_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")

    # Line chart
    chart = LineChart()
    chart.title  = "Monthly Revenue"
    chart.y_axis.title = "Revenue ($)"
    chart.x_axis.title = "Month"
    chart.style = 10
    chart.width  = 24
    chart.height = 14

    data_ref = Reference(ws, min_col=2, min_row=2, max_row=len(monthly)+2)
    cats_ref = Reference(ws, min_col=1, min_row=3, max_row=len(monthly)+2)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    ws.add_chart(chart, f"A{next_row + 1}")


def _build_segment_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("Customer Segments")
    ws.sheet_view.showGridLines = False

    seg_df = (
        df.groupby("segment")
        .agg(
            unique_customers=("customer_id","nunique"),
            total_orders    =("order_id",   "nunique"),
            total_revenue   =("sales",      "sum"),
            avg_order_value =("sales",      "mean"),
            total_profit    =("profit",     "sum"),
        )
        .round(2)
        .reset_index()
        .sort_values("total_revenue", ascending=False)
    )

    _write_df_to_sheet(ws, seg_df, start_row=2, header_color=MID_BLUE)

    ws.merge_cells("A1:F1")
    h = ws["A1"]
    h.value     = "Revenue & Orders by Customer Segment"
    h.font      = Font(bold=True, size=13, color=WHITE, name="Calibri")
    h.fill      = _header_fill(DARK_BLUE)
    h.alignment = Alignment(horizontal="center", vertical="center")


# ── Main entry point ──────────────────────────────────────────────────────────

def generate_report(csv_path: str = CLEAN_CSV,
                    output_dir: str = REPORTS_OUTPUT_DIR) -> str:
    """Build the Excel report and return its file path."""
    print("\n[report_generator] Loading data …")
    df = pd.read_csv(csv_path, low_memory=False)

    wb = Workbook()
    # Remove default empty sheet
    wb.remove(wb.active)

    print("[report_generator] Building sheets …")
    _build_summary_sheet(wb, df)
    _build_revenue_by_region(wb, df)
    _build_top_products(wb, df)
    _build_category_sheet(wb, df)
    _build_monthly_trend(wb, df)
    _build_segment_sheet(wb, df)

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_path  = os.path.join(output_dir, f"sales_kpi_report_{timestamp}.xlsx")
    wb.save(out_path)
    print(f"[report_generator] ✅ Report saved → {out_path}\n")
    return out_path


if __name__ == "__main__":
    generate_report()
